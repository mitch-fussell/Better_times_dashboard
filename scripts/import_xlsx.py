"""One-time import of the BetterTimes spreadsheet into seed SQL.

Reads the hand-coloured check-in grid and emits INSERT statements for the
`clients` and `check_ins` tables. Colour -> type mapping:
  green  FF00B050 -> proactive  (we reached out)
  yellow FFFFFF00 -> onboarding
  red    FFFF0000 -> reactive   (client reported an issue)

Grid layout (sheet "May 2026"):
  col B, rows 4..47   = client names
  cols D..AH (4..34)  = May 2026, day = col-3
  cols AI..BL (35..64)= June 2026, day = col-34
"""
import sys
import openpyxl

SRC = sys.argv[1] if len(sys.argv) > 1 else "/tmp/bettertimes.xlsx"
COLOR_TYPE = {"FF00B050": "proactive", "FFFFFF00": "onboarding", "FFFF0000": "reactive"}


def cell_color(cell):
    f = cell.fill
    if not f or not f.patternType:
        return None
    fg = f.fgColor
    return fg.rgb if fg and fg.type == "rgb" else None


def date_for_col(c):
    if 4 <= c <= 34:
        return f"2026-05-{c - 3:02d}"
    if 35 <= c <= 64:
        return f"2026-06-{c - 34:02d}"
    return None


def q(s):
    return "'" + s.replace("'", "''") + "'"


wb = openpyxl.load_workbook(SRC)
ws = wb["May 2026"]

clients = []          # (row, name)
checkins = []         # (name, date, type)
for r in range(4, 48):
    name = ws.cell(r, 2).value
    if not name or not str(name).strip():
        continue
    name = str(name).strip()
    clients.append(name)
    for c in range(4, 65):
        t = COLOR_TYPE.get(cell_color(ws.cell(r, c)))
        if not t:
            continue
        d = date_for_col(c)
        if d:
            checkins.append((name, d, t))

# latest check-in type per client -> status
latest = {}
for name, d, t in checkins:
    if name not in latest or d > latest[name][0]:
        latest[name] = (d, t)

lines = ["begin;"]
for name in clients:
    status = "onboarding" if latest.get(name, (None, ""))[1] == "onboarding" else "active"
    lines.append(
        f"insert into clients (name, status, cadence_days) values ({q(name)}, '{status}', 7) "
        f"on conflict (name) do nothing;"
    )
for name, d, t in checkins:
    lines.append(
        "insert into check_ins (client_id, occurred_on, type) "
        f"select id, '{d}', '{t}' from clients where name = {q(name)};"
    )
lines.append("commit;")

out = "\n".join(lines)
with open("/tmp/seed.sql", "w") as fh:
    fh.write(out)

print(f"clients={len(clients)} check_ins={len(checkins)}")
by_type = {}
for _, _, t in checkins:
    by_type[t] = by_type.get(t, 0) + 1
print("by_type=", by_type)
print("wrote /tmp/seed.sql")
